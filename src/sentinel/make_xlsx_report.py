"""Detailed Excel report with all per-sample fields and a column legend.

Reads `<PROJ>/results/per_sample_report.tsv` and writes
`<PROJ>/results/report.xlsx`.
"""
from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


PROJ = Path(os.environ.get("PROJ", "."))
RESULTS_DIR = PROJ / "results"
OUT = RESULTS_DIR / "report.xlsx"


COLUMN_LEGEND = [
    ("sample_id", "Sample identifier as supplied in the sample sheet."),
    ("well", "Plate well coordinate (e.g. A1, H12) when provided in the sample sheet."),
    ("sample_type", "control / reference_fresh / reference_ffpe / mix / clinical (defaults to clinical)."),
    ("verdict", "PASS / WARN / FAIL after applying sample-type-aware thresholds."),
    ("identity_match", "TRUE when the genetics match the expected identity; FALSE marks a swap or heavy contamination."),
    ("concordance_to_expected", "Genetic similarity (0-1) to the sample's expected identity row."),
    ("best_match_sample_id", "Closest-matching sample on the run. For a clean sample this equals the expected identity."),
    ("best_match_concordance", "Genetic similarity to the closest-matching sample on the run."),
    ("top_score_homalt", "Headline cross-sample contamination score; approximates the contamination fraction."),
    ("top_donor_sample_id", "Best single guess at the contamination source for this recipient."),
    ("top_n_informative_homalt", "Informative sites supporting the top donor inference."),
    ("homalt_deflation", "Drop in alt-VAF at the sample's own hom-alt sites; rises with foreign-read mixing."),
    ("homalt_vaf", "Mean alt-VAF at the sample's own hom-alt sites; expected near 1 in clean samples."),
    ("n_homalt_sites", "Number of called hom-alt sites used by the deflation metric."),
    ("frac_pairs_3plus_haps", "Fraction of phaseable site pairs supporting more than two haplotypes."),
    ("mean_minor_hap_fraction", "Mean within-pair fraction of the minor haplotype."),
    ("n_haplotype_pairs_eval", "Number of phaseable site pairs evaluated."),
    ("vaf_tail_fraction", "Fraction of called sites with mid-range alt-VAF (anomaly indicator)."),
    ("n_consensus_reads", "Consensus reads sampled from the BAM."),
    ("median_molecules_per_site", "Median molecule support per scout site."),
    ("dup_factor", "Raw / consensus duplication factor."),
    ("n_called_sites", "Total number of sites called for this sample."),
    ("background_alt_vaf", "Per-sample baseline alt-VAF at hom-ref sites (sequencing noise floor)."),
    ("top_donor_tied_set", "All candidate donors tied within a tolerance band of the top score."),
]


def _write_legend(ws):
    ws.cell(row=1, column=1, value="Column").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Description").font = Font(bold=True)
    for i, (col, desc) in enumerate(COLUMN_LEGEND, start=2):
        ws.cell(row=i, column=1, value=col).font = Font(name="Consolas")
        c = ws.cell(row=i, column=2, value=desc)
        c.alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 110


def main():
    in_tsv = RESULTS_DIR / "per_sample_report.tsv"
    if not in_tsv.exists():
        raise FileNotFoundError(in_tsv)
    df = pd.read_csv(in_tsv, sep="\t")

    wb = Workbook()
    ws = wb.active
    ws.title = "Per-sample report"

    ws["A1"] = "Sentinel - per-sample report"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:F1")

    header_row = 3
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_idx, row in enumerate(rows, start=header_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == header_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
            else:
                col_name = df.columns[c_idx - 1] if c_idx - 1 < len(df.columns) else None
                if col_name == "verdict":
                    if value == "PASS":
                        cell.fill = PatternFill("solid", fgColor="DCFCE7")
                    elif value == "WARN":
                        cell.fill = PatternFill("solid", fgColor="FEF3C7")
                    elif value == "FAIL":
                        cell.fill = PatternFill("solid", fgColor="FECACA")

    legend_ws = wb.create_sheet("Column legend")
    _write_legend(legend_ws)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
